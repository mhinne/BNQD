# -*- coding: utf-8 -*-
"""
Created on Tue Sep 10 12:30:28 2019

@author: u341138
"""

import sys

import GPy
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.path as mplPath
import matplotlib.tri as mtri
from matplotlib import cm

import os
# for plotting the Basemap figure, I had to add the following environment variable.
os.environ['PROJ_LIB'] = r'C:\\Users\\u341138\\AppData\\Local\\Continuum\\anaconda3\\pkgs\\proj4-5.2.0-h6538335_1006\\Library\\share'


from mpl_toolkits.basemap import Basemap
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.collections import PolyCollection
from mpl_toolkits.axes_grid1 import ImageGrid
#from mpl_toolkits import mplot3d

from pykml import parser
# NB: had to change parser.py, see https://umar-yusuf.blogspot.com/2018/06/unable-to-parse-kml-file-in-python-3.html
from openpyxl import load_workbook

import csv
import warnings; warnings.simplefilter('ignore')

import importlib
sys.path.append('../')
import BNQD
importlib.reload(BNQD)

print("GPy version:      {}".format(GPy.__version__))
print("BNQD version:     {}".format(BNQD.__version__))


plt.close('all')

print('Analysis of Dutch voting behaviour')

NL_ll_lon, NL_ll_lat, NL_ur_lon, NL_ur_lat = 3.0, 50.5, 7.5, 54.0

map_resolution = 'h'

          
def read_data(votesfile='..\\datasets\\Dutch parliament elections\\Gemeente_uitslagen_TK2017.csv', 
              geofile='..\\datasets\\Dutch parliament elections\\Gemeenten_2019_locations.csv'): 
    print('Reading data')
    
    geo_data = dict()
    
    with open(geofile, 'r') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')
        for row in reader:
            geo_data[row['NAAM']] = {'Provincie': row['Provincie'], 
                    'x': float(row['Lon'].replace(',', '.')), 
                    'y': float(row['Lat'].replace(',', '.'))}
    
    data = dict()
    
    with open(votesfile, 'r') as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')
        for row in reader:
            label = row['Gemeente']
            if label in geo_data:
                data[label] = {'VVD': int(row['VVD']),
                                 'CDA': int(row['CDA']),
                                 'PVV': int(row['PVV']),
                                 'D66': int(row['D66']),
                                 'SP': int(row['SP']),
                                 'GROENLINKS': int(row['GROENLINKS']),
                                 'PvdA': int(row['PvdA']),
                                 'CU': int(row['CU']),
                                 '50PLUS': int(row['50PLUS']),
                                 'PvdD': int(row['PvdD']),
                                 'SGP': int(row['SGP']),
                                 'FvD': int(row['FvD']),
                                 'Provincie': geo_data[label]['Provincie'],
                                 'Lon': geo_data[label]['x'],
                                 'Lat': geo_data[label]['y']
                                 }
    return data

def read_province_data(provincefile='..\\datasets\\Dutch parliament elections\\Provincies Nederand KML _ LocalFocus.xlsx'):
        
    def get_col_by_header(ws, value, row=1):
        for i, col in enumerate(ws.iter_cols()):
            if col[row].value == value:
                return i+1
        return -1
        
    wb_source = load_workbook(provincefile)    
    wb_source.active = wb_source['Blad1']    
    ws_source = wb_source.active
    
    provinces = dict()
    startsAt = 2
    
    province_polygon_col = get_col_by_header(ws_source, 'KML', row=0)
    
    for i, row in enumerate(ws_source.iter_rows(min_row=startsAt, max_row=13)):        
        province_name = row[0].value  
        province_surface = ws_source.cell(row = i+startsAt, 
                                          column = province_polygon_col).value
        provinces[province_name] = province_surface        
    
    province_polygons = dict()
    
    for province in provinces.keys():
        province_polygons[province] = list()
        for subpoly in parser.fromstring(provinces[province]).getchildren():
            poly_coords = str(subpoly.coordinates).split(' ')
            poly_coords_list = [pc.split(',') for pc in poly_coords]
            poly_x = [float(i[0]) for i in poly_coords_list]
            poly_y = [float(i[1]) for i in poly_coords_list]
            province_polygons[province].append((poly_x, poly_y))
            
    return province_polygons 
    
def total_vote_count(data):
    total_votes = {label:  data[label]['VVD'] + 
                           data[label]['CDA'] + 
                           data[label]['PVV'] + 
                           data[label]['D66'] + 
                           data[label]['SP'] + 
                           data[label]['GROENLINKS'] + 
                           data[label]['PvdA'] + 
                           data[label]['50PLUS'] + 
                           data[label]['CU'] + 
                           data[label]['PvdD'] + 
                           data[label]['SGP'] +
                           data[label]['FvD'] for label in data.keys()}
    return total_votes

def flat_plot(x, y, z, ix, cmap=cm.viridis, marker_alpha=0.7, label1='Set 1', 
              label2='Set 2', title=None):
    
    fig = plt.figure(figsize=(9,9))
    m = Basemap(projection='merc', llcrnrlat=NL_ll_lat, urcrnrlat=NL_ur_lat, 
                llcrnrlon=NL_ll_lon, urcrnrlon=NL_ur_lon, resolution=map_resolution)
    m.drawmapboundary(fill_color='#d8eaed')
    m.drawcountries(linewidth=1.0)
    m.drawcoastlines()
    m.fillcontinents(color='#f0efe9', lake_color='#d8eaed')
    m.scatter(x[ix], y[ix], 
              c=z[ix], alpha=marker_alpha, cmap=cmap, marker='s', zorder=10, 
              latlon=True, label=label1, edgecolor='k')
    sc = m.scatter(x[np.logical_not(ix)], y[np.logical_not(ix)], 
              s=30, c=z[np.logical_not(ix)], alpha=marker_alpha, cmap=cmap, 
              marker='o', zorder=10, latlon=True, label=label2, edgecolor='k')
    if title:
        plt.title(title)
    plt.legend(loc='upper left')
    cb = plt.colorbar(sc)
    cb.set_label('Fraction of votes to populist parties')
    plt.show()
    return fig, m


def full_plot(x, y, z, ix, cmap=cm.summer, marker_alpha=0.7, label1='Set 1', label2='Set 2', title=None):
    
    m = Basemap(projection='merc', llcrnrlat=NL_ll_lat, urcrnrlat=NL_ur_lat, 
                llcrnrlon=NL_ll_lon, urcrnrlon=NL_ur_lon, 
                resolution=map_resolution)
    fig = plt.figure(figsize=(9,9))
    ax = Axes3D(fig)
    ax.azim = 270
    ax.elev = 90
    ax.dist = 10
    
    ax.add_collection3d(m.drawcountries())    
    
    # replaces fillcontinents
    polys = []
    for polygon in m.landpolygons:
        polys.append(polygon.get_coords())    
    
    lc = PolyCollection(polys, edgecolor='black',
                        facecolor='#f0efe9', closed=False, linewidth=1.0)
    
    ax.add_collection3d(lc)
    m.scatter(x[ix], y[ix], z[ix],
              c=z[ix], alpha=marker_alpha, cmap=cmap, marker='s', zorder=10, 
              latlon=True, label=label1, edgecolor='k')
    m.scatter(x[np.logical_not(ix)], y[np.logical_not(ix)], 
                z[np.logical_not(ix)], s=30, c=z[np.logical_not(ix)], 
                alpha=marker_alpha, cmap=cmap, marker='o', zorder=10, 
                latlon=True, label=label2, edgecolor='k')
    if title:
        plt.title(title)
    plt.legend(loc='best')
    plt.show
    return fig, m
    
def flat_province(fig, m, province):
    ax = fig.gca()    
    for landmass in province:
        x, y = m(landmass[0], landmass[1])
        ax.plot(x, y, linewidth=2, color='k')
        
def flat_draw_border(ax, m, a, b, xmin=NL_ll_lon, xmax=NL_ur_lon, label='Line'):
    x = np.linspace(xmin, xmax, 100)
    y = a*x + b
    
    xx, yy = m(x, y)
    ax.plot(xx, yy, linestyle='--', linewidth=3.0, label=label, color='k', 
            zorder=10)
    
def full_draw_border(fig, m, a, b, xmin=NL_ll_lon, xmax=NL_ur_lon, label='Line'):
    x = np.linspace(xmin, xmax, num=100)
    y = a*x + b
    ax = fig.gca()
    xx, yy = m(x, y)
    ax.plot3D(xx, yy, 1e-5*np.ones((100)))

def full_province(fig, m, province):
    ax = fig.gca(projection='3d')
    province_polys = []
    for polygon in province:
        px, py = m(polygon[0], polygon[1])
        province_polys.append(np.column_stack((px, py)))
    
    pplc = PolyCollection(province_polys, edgecolor='black', 
                          facecolor='#ffffff', closed=False)    
    ax.add_collection3d(pplc, zs=1e-4)      

def inProvince(x, province):
    # check whether point is in province
    for polygon in province:
        bbPath = mplPath.Path(np.column_stack((polygon[0], polygon[1])))
        if bbPath.contains_point(x):
            return True
    return False

def inCountry(X, provinces):
    # check whether point is in the Netherlands    
    valid_points = list()
    n = X.shape[0]
    
    for i in range(n):
        x = X[i,:]
    
        for k, v in provinces.items():
            if inProvince(x, v):
                valid_points.append(x)
    return np.array(valid_points)

def belowPhantomBorder(x, a, b):
    y_border = a*x[0]+b
    return x[1] < y_border

def make_triangles(x, y, threshold):
    tri = mtri.Triangulation(x, y)
    triangles = tri.triangles
    mask = list()
    for triangle in triangles:    
        u, v, w = triangle
        d0 = np.sqrt( (x[u] - x[w]) **2 + (y[u] - y[v])**2 )
        d1 = np.sqrt( (x[v] - x[v]) **2 + (y[v] - y[w])**2 )
        d2 = np.sqrt( (x[w] - x[u]) **2 + (y[w] - y[u])**2 )
        max_edge = np.max([d0, d1, d2])
        if max_edge > threshold:
            mask.append(True)
        else:
            mask.append(False)
    tri.set_mask(mask) 
    return tri

data = read_data()
X = np.column_stack((np.array([data[label]['Lon'] for label in data.keys()]), np.array([data[label]['Lat'] for label in data.keys()])))
n, D = X.shape
total_votes = total_vote_count(data)


print('Analyze populism')
a = 0.15
b = 51
border_xmin, border_xmax = (3.7, 6.75)
y = np.array([float(data[label]['PVV'] + \
                    data[label]['SP'] + \
                    data[label]['50PLUS'] + \
                    data[label]['FvD'] ) / total_votes[label] 
              for label in data.keys()])
labelFunc = lambda x: [belowPhantomBorder(i, a, b) for i in x]

  
x_res = 200
y_res = 200
provinces = read_province_data()
xx, yy = np.meshgrid(np.linspace(NL_ll_lon, NL_ur_lon, x_res), np.linspace(NL_ll_lat, NL_ur_lat, y_res))
X_pred = np.column_stack((xx.flatten(), yy.flatten()))
X_pred = inCountry(X_pred, provinces)

bn = 100
border_x1 = np.linspace(border_xmin, border_xmax, num=bn)
border_x2 = a*border_x1 + b
border_pred = np.column_stack((border_x1, border_x2))


num_restarts    = 50
kernel_names    = ['Linear', 'Exponential', 'RBF', 'Matern32']
linear_kernel   = GPy.kern.Linear(D) + GPy.kern.Bias(D)
exp_kernel      = GPy.kern.Exponential(D)
Matern32_kernel = GPy.kern.Matern32(D)
RBF_kernel      = GPy.kern.RBF(D)
kernels         = [linear_kernel, exp_kernel, RBF_kernel, Matern32_kernel]

kernel_dict     = dict(zip(kernel_names, kernels))

opts = dict()
opts['num_restarts'] = num_restarts
opts['mode'] = 'BIC'
opts['verbose'] = False

Total = BNQD.BnpQedAnalysis(X, y, kernel_dict, labelFunc, b=border_pred, 
                            opts=opts)
Total.train()
Total.pretty_print(verbose=True)
    

# effect size plot
color_disc = '#cc7d21'

# How to plot BMA here?
fig, axes = plt.subplots(nrows=1, ncols=len(kernels), figsize=(18,4), 
                         sharex=True, sharey=True)    
for i, (k, v) in zip(range(len(kernels)), kernel_dict.items()): 

    preds = Total.results[k].summary(b=border_pred)
    es_mean = np.array([preds['es_disc_stats'][i][0] for i in range(bn)])
    es_std  = np.array([preds['es_disc_stats'][i][1] for i in range(bn)])
                
    axes[i].plot(border_x1, es_mean, label='Effect size', linewidth=1.0, 
                 linestyle='--', color=color_disc)
    axes[i].fill_between(border_x1, np.squeeze(es_mean + 0.5*es_std), 
                         np.squeeze(es_mean - 0.5*es_std), 
                         color=color_disc, alpha=0.3)
    axes[i].annotate('log BF = {:0.1f}'.format(preds['logbayesfactor']), 
                     (3.8, 0.12))
    axes[i].autoscale(enable=True, axis='x', tight=True)
    axes[i].yaxis.tick_right()
    axes[i].set_ylabel('Vote fraction')
    axes[i].axhline(y=0.0, linestyle='--', color='k')        

axes[1].set_xlabel('Phantom border (longitude; degrees)')
axes[1].set_ylim([-0.1, 0.15])
plt.suptitle('2D effect size')

total_bma = Total.get_total_log_Bayes_factor(verbose=True)

for kernel in kernel_names:
    summ = Total.results[kernel].summary(b=border_pred)
    print('{:s}, log BF = {:0.1f}, log p(D|M_C) = {:0.3f}, log p(D|M_D) = {:0.3f}'.format(kernel, 
          summ['logbayesfactor'],
          summ['evidence']['mc'],
          summ['evidence']['md']))


# geo map plots

cmap = cm.viridis

for k in ['Linear', 'Exponential']:
    # Plotting
    gprdd = Total.results[k]
    stats = gprdd.summary(b=border_pred)
    y_pred_c_mean, _ = gprdd.CModel.predict(X_pred)
    y_pred_d = gprdd.DModel.predict(X_pred)
    zz = np.squeeze(y_pred_c_mean)        
    
    fig = plt.figure(figsize=(18,8))        
    grid = ImageGrid(fig, 111,          
             nrows_ncols=(1,2),
             axes_pad=0.15,
             share_all=True,
             cbar_location="right",
             cbar_mode="single",
             cbar_size="7%",
             cbar_pad=0.15,
             )
    
    for ax in grid:            
        m = Basemap(projection='merc', 
                    llcrnrlat=NL_ll_lat, 
                    urcrnrlat=NL_ur_lat, 
                    llcrnrlon=NL_ll_lon, 
                    urcrnrlon=NL_ur_lon, 
                    ax=ax, 
                    resolution='h')
        m.drawmapboundary(fill_color='#d8eaed')
        m.drawcountries(linewidth=1.0, zorder=11)
        m.drawcoastlines(zorder=11)
        m.fillcontinents(color='#f0efe9', lake_color='#d8eaed')
                         
    grid[0].set_title('Continuous model')
    grid[0].set_ylabel(k)
    grid[1].set_title('Discontinuous model')
                         
    xx, yy = m(X_pred[:,0], X_pred[:,1])            
    
    tcf = grid[0].tricontourf(make_triangles(xx, yy, threshold=18000), 
              zz, levels=np.linspace(0.1, 0.6, num=30), zorder=10, cmap=cmap)
    
    cb = grid[1].cax.colorbar(tcf)
    cb.set_label_text('Fraction of votes for populist parties')
    plt.suptitle('{:s} kernel, log Bayes factor = {:0.1f}'.format(k, stats['logbayesfactor']))
    print('{:s} kernel, log Bayes factor = {:0.1f}'.format(k, stats['logbayesfactor']))
    
    for i, X_pred_split in enumerate([X_pred[labelFunc(X_pred),:], X_pred[np.logical_not(labelFunc(X_pred)),:]]):    
        zz = np.squeeze(y_pred_d[i][0])
        
        x1, x2 = m(X_pred_split[:,0], X_pred_split[:,1])
        tcf = grid[1].tricontourf(make_triangles(x1, x2, threshold=18000), 
                  zz, levels=np.linspace(0.1, 0.6, num=30), zorder=10, cmap=cmap)
        
    for ax in grid:
        XX, YY = m(X[:,0], X[:,1])
        ax.scatter(XX, YY, 
          c=y, alpha=1.0, cmap=cmap, marker='o', zorder=10, label='test', edgecolor='k')
        flat_draw_border(ax, m, a=a, b=b, label='Phantom border')

